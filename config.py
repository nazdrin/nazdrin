
token = ('2007585933:AAHWk-Z9fq8Fa5iyojR500w5_vurQ5Qxz5Q')
token_2 = ('1185190149:AAGaCD5jNEXgHZ3HYUpMJlb7K---mZQ8SR0')
import openpyxl
from data_base import sqlite_db
import json




#bot = Bot(token = '2119047086:AAG2tUZmRULNasX8Rkb1tTUb4uEDwkQ640k') aiogranmstury bot
# admin = -736955962
# button_dict = {'Товари💊':                  {'🔎Товар не відображається на сайті': [answers.no_goods,answers.hours_6],'📥Додати товар на сайт': [answers.add_goods,answers.hours_24],
#                                             '🪢Змінити прив"язку товара🪢' : [answers.change_binding,answers.hours_6] },
#               'Відображення даних аптек🏥': {'🏥Додати нові аптеки':[answers.add_pharmacies,answers.hours_24],'🔎Аптека не відображається':[answers.show,answers.hours_1],
#                                             '📜☎ Змінити графік роботи, номер телефона️':[answers.schedule,answers.hours_24],'🗺Змінити точку на карті':[answers.point_map,answers.hours_24],
#                                              '🆕Змінити назву аптеки' : [answers.change_name_store,answers.hours_24],
#                                              '🚫Відключити аптеку':[answers.cancel_pharmacy,answers.hours_1],'❌Відключити мережу':[answers.cancel_chain,answers.hours_1] },
#                'Договори,рахунки,акти📑': {'📜Договори': [answers.contract,answers.hours_48], '🧾Рахунки': [answers.bills,answers.hours_6],
#                                            '📇Акти' : [answers.acts,answers.hours_6],'🏃Зміна контактної особи' :[answers.change_contact,answers.hours_24]},
#                'Тех.збій🛠' :             { '📵Особистий кабінет' : [answers.personal,answers.hours_1],'⛔️Замовлення': [answers.orders,answers.hours_1],'⭕️Залишки': [answers.stock,answers.hours_2]},
#                'Звіти📈' :                { '🪢Товари без прив"язки' : [answers.no_con,answers.hours_24], '📈Якість' : [answers.quality,answers.hours_24], '🚫Заблоковані товари': [answers.block_goods,answers.hours_6],
#                                           '🗺Оточення' : [answers.around,answers.hours_24], '💰Фінансовий' : [answers.finance,answers.hours_24]}}


# with open('number.txt') as json_file:  # Заполняем словарь с ID пользователей из Json при перезапуске бота
#     number_request = json.load(json_file)
wb = openpyxl.load_workbook('Pharma.xlsx')  # Заполняем словарь pharmacies_chain из єксель при перезапуске бота
sheet = wb.active
rows = sheet.max_row

for i in range(1, rows + 1):
    a = sheet.cell(row=i, column=1)
    b = sheet.cell(row=i, column=2)
    c = sheet.cell(row=i, column=3)


    code = a.value
    name = b.value
    manager = c.value
    sqlite_db.sql_insert_new_chain(code, name, manager)



with open('data.txt') as json_file:  # Заполняем словарь с ID пользователей из Json при перезапуске бота
    users = json.load(json_file)

for key, value in users.items():
    users_name = ''
    sqlite_db.sql_new_user(key, value,users_name)

